import openpyxl
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from lxml import etree
from RCA import Config as cf
import re


class Utility:

    def get_classification_list(workbookpath):
        workbook = openpyxl.load_workbook(workbookpath)

        sheet = workbook['Sheet1']

        col = sheet['C']

        datalist = []

        for x in range(len(col)):
            value = col[x].value
            value = str(value)
            datalist.append(value)

        return datalist

    def get_col_list(workbookpath, sheetname, col_code):
        workbook = openpyxl.load_workbook(workbookpath)

        sheet = workbook[sheetname]

        col = sheet[col_code]

        datalist = []

        for x in range(len(col)):
            if x > 0:
                value = col[x].value
                if value is None:
                    pass
                else:
                    value = str(value)
                    datalist.append(value)

        return datalist

    def get_province(workbookpath, sheetname):
        workbook = openpyxl.load_workbook(workbookpath)

        sheet = workbook[sheetname]

        col = sheet['A']

        province_name = ''
        if col[1]:
            province_name = str(col[1].value)
        else:
            print('读取省份名称失败！')
        return province_name

    def write_temp_pagelinklist(expected_crawl_links_list, writefilepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '地区名称'
        ws['B1'] = '分类词汇'
        ws['C1'] = '生成的初始链接'

        for regions in range(len(expected_crawl_links_list)):
            ws.append(expected_crawl_links_list[regions].split('$$$'))
        wb.save(writefilepath)

    def write_temp_allpagelinklist(self, expected_crawling_links, writefilepath):
        innerlinks = []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '地区名称'
        ws['B1'] = '分类词汇'
        ws['C1'] = '生成的全部链接'
        config = cf.Config()
        for regions in range(len(expected_crawling_links)):
            tempsrn = expected_crawling_links[regions].split('$$$')[0]
            tempsc = expected_crawling_links[regions].split('$$$')[1]
            seedlink = expected_crawling_links[regions].split('$$$')[2]
            tempinnerlinks = Utility.get_links(self, seedlink, config.reCrawlNumber)

            for i in range(len(tempinnerlinks)):
                innerlinks.append(tempsrn + '$$$' + tempsc + '$$$' + tempinnerlinks[i])
            print('processing——' + tempsrn + tempsc + str(regions))

        grouplinks = list(set(innerlinks))
        grouplinks.sort(key=innerlinks.index)

        for item in range(len(grouplinks)):
            ws.append(grouplinks[item].split('$$$'))
        wb.save(writefilepath)

    def get_links(self, url, re_crawl_num):
        chrome_options = Options()

        ua = UserAgent()
        header = {
            'User-Agent': ua.random
        }

        chrome_options.add_argument('user-agent="%s"' % header)

        # 设置打开浏览器不加载图片
        chrome_options.add_argument('–blink - settings = imagesEnabled = false')
        # # 不加载图片,加快访问速度
        chrome_options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
        # 此步骤很重要，设置为开发者模式，防止被各大网站识别出来使用了Selenium
        chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])

        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')

        driver = webdriver.Chrome(options=chrome_options)
        driver.set_page_load_timeout(20)

        try:
            driver.get(url)

            html = driver.page_source

            driver.quit()

            xp_data = etree.HTML(html)

            links = xp_data.xpath('//h3[@class="t"]/a/@href|//h3[@class="t c-gap-bottom-small"]/a/@href')
        except:
            links = None
            if re_crawl_num >= 0:
                self.get_links(url, re_crawl_num - 1)
        return links

    def read_txt_to_list(path):
        wordlist = []
        try:
            file = open(path, 'r', encoding='utf8')

            for wordinline in file.readlines():
                wordlist.append(wordinline.strip())
        except:
            pass
        return wordlist

    def delete_district(word):
        temp_return_word = ''
        tword = word.replace(' ', '')  # 去空格
        result = re.findall(
            r'[\u4e00-\u9fa5]{1,4}省|[\u4e00-\u9fa5]{1,4}市|[\u4e00-\u9fa5]{1,4}区|[\u4e00-\u9fa5]{1,4}县|[\u4e00-\u9fa5]{1,4}乡|[\u4e00-\u9fa5]{1,4}镇|[\u4e00-\u9fa5]{1,9}州|[\u4e00-\u9fa5]{0,4}街道|[\u4e00-\u9fa5]{1,4}村|[\u4e00-\u9fa5]{1,4}旗|[\u4e00-\u9fa5]{1,4}屯|[\u4e00-\u9fa5]{1,4}庄',tword)  # 分别匹配模式

        if len(result) > 0:
            pass
        else:
            # 不包含行政区划，带关键词“省”......
            temp_return_word = word

        return temp_return_word

    def delete_gov(word):
        temp_return_word = ''
        tword = word.replace(' ', '')  # 去空格
        result = re.findall(r'[\u4e00-\u9fa5]{1,4}部|[\u4e00-\u9fa5]{1,6}局|[\u4e00-\u9fa5]{1,4}工会|[\u4e00-\u9fa5]{0,5}学校|[\u4e00-\u9fa5]{1,6}路|[\u4e00-\u9fa5]{0,4}机关|[\u4e00-\u9fa5]{0,4}公署|[\u4e00-\u9fa5]{0,4}政府|[\u4e00-\u9fa5]{0,4}支队|[\u4e00-\u9fa5]{1,3}电|[\u4e00-\u9fa5]{0,4}处|[\u4e00-\u9fa5]{0,4}部门|[\u4e00-\u9fa5]{0,6}集团|[\u4e00-\u9fa5]{1,4}厅|[\u4e00-\u9fa5]{0,4}产品|[\u4e00-\u9fa5]{0,4}设施|[\u4e00-\u9fa5]{0,4}功能|[\u4e00-\u9fa5]{0,6}公司|[\u4e00-\u9fa5]{1,6}会|[\u4e00-\u9fa5]{1,6}院|[\u4e00-\u9fa5]{1,6}银行|[\u4e00-\u9fa5]{0,4}委|[\u4e00-\u9fa5]{0,4}室|[\u4e00-\u9fa5]{0,4}队|[\u4e00-\u9fa5]{0,5}业|[\u4e00-\u9fa5]{1,5}所|[\u4e00-\u9fa5]{1,5}方|[\u4e00-\u9fa5]{1,5}社|[\u4e00-\u9fa5]{0,4}合同|[\u4e00-\u9fa5]{0,4}保险|[\u4e00-\u9fa5]{0,4}中心|[\u4e00-\u9fa5]{0,4}管理',tword)  # 分别匹配模式

        if len(result) > 0:
            pass
        else:
            # 不包含行政区划，带关键词“省”......
            temp_return_word = word

        return temp_return_word

    def delete_usual(word):
        temp_return_word = ''
        tword = word.replace(' ', '')  # 去空格
        result = re.findall(r'[\u4e00-\u9fa5]{0,4}资源|[\u4e00-\u9fa5]{0,6}中学',tword)  # 分别匹配模式

        if len(result) > 0:
            pass
        else:
            # 不包含usual word
            temp_return_word = word

        return temp_return_word

    def delete_common_placename(word, placenamelist):
        temp_return_word = ''
        if placenamelist:
            if word in placenamelist:
                pass
            else:
                # 不在常用地名中t
                temp_return_word = word
        else:
            print('需要载入常用省市名称词表！')

        return temp_return_word

    def delete_common_miniplacename(word, placenamelist):   #去除最小粒度村名
        temp_return_word = ''
        if placenamelist:
            if word in placenamelist:
                pass
            else:
                # 不在常用村名中t
                temp_return_word = word
        else:
            print('需要载入常用村名称词表！')

        return temp_return_word
